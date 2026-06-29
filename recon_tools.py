#!/usr/bin/env python3
"""
recon_tools.py — deterministyczny silnik skilla competitor-recon.

Analizuje eksport widoczności konkurenta (Senuto / Ahrefs / SEMrush) i typuje,
które STRONY konkurenta są najważniejsze do odtworzenia, by przejąć ich ruch.

Podział pracy (świadomie):
  • TEN SKRYPT robi MECHANIKĘ: parsowanie eksportu, agregację keyword→URL,
    regułową klasyfikację fraz (brand / generyk-podejrzany / komercyjne / lokalne /
    informacyjne / obcojęzyczne), wybór kandydatów na "main keyword", scoring i tiering.
  • CLAUDE robi OSĄD SEMANTYCZNY (krok między `ingest` a `score`): potwierdza, które
    "generyki-podejrzane" są realnie bezwartościowe/ambiwalentne (np. "landing",
    "biogram"), nadaje finalny typ strony, finalny main keyword i dopasowanie
    biznesowe (relevance) do NASZEJ domeny.

Komendy:
  ingest <plik.xlsx|csv> [--domain kamanmarketing.pl] [--our-domain double-digital.pl]
         [--top 40] [--out DIR]
      → DIR/keywords.csv, DIR/pages.json, DIR/worksheet.md (do osądu Claude'a)

  score  --in DIR --judgments DIR/judgments.json [--out DIR]
      → DIR/report.md, DIR/priorities.csv

Uruchamiać przez /usr/local/bin/python3 (popsuty venv w repo).
"""
import sys, os, csv, json, re, argparse, math
from collections import defaultdict

# ───────────────────────── konfiguracja / wagi (tunable) ─────────────────────────

# wagi composite Opportunity Score
W_TRAFFIC, W_BREADTH, W_COMMERCIAL = 0.60, 0.20, 0.20
KD_PENALTY = 0.30          # maks. udział kary za trudność
CPC_CAP = 10.0             # CPC powyżej → pełny sygnał komercyjny (po winsoryzacji)
CPC_WINSOR = 60.0          # twardy cap CPC: powyżej = niemal zawsze błąd danych Senuto (995/202/128…)
COMMERCIAL_CPC = 3.0       # CPC >= → traktuj frazę jako komercyjną
GENERIC_VOL = 2000         # 1-token + vol>= → podejrzenie generyka
RELEVANCE_SKIP = 0.30      # relevance < → SKIP (niskie dopasowanie)
RELEVANCE_FLOOR_P3 = 0.50  # relevance < → strona NIE może być wyżej niż P3 (twardy floor na traffic-bait)
# tiery WZGLĘDNE wśród osądzonych kwalifikujących się stron (percentyl po score):
TIER_P1_PCT = 0.20         # top 20% → P1
TIER_P2_PCT = 0.55         # następne do 55% → P2; reszta → P3
OPEN_GOAL_MIN = 1500       # latent_potential >= → strona to "open goal" (popyt jest, konkurent słabo rankuje)
WEAK_POS_MIN = 8           # pozycja konkurenta >= → fraza realnie niezdobyta (poniżej połowy 1. strony)
LATENT_MIN_VOL = 50        # licz do latent tylko frazy z wolumenem >= (odsiewa drobny ogon)
LATENT_W = 0.12            # ile latentu wchodzi do "opportunity" (popyt do odzyskania ≈ vol × CTR na dobrej pozycji)

# intent → waga frazy w "qualified traffic" (gdy brak osądu Claude'a)
INTENT_WEIGHT = {
    "brand": 0.0, "foreign": 0.0, "generic_suspect": 0.25,
    "commercial": 1.0, "local": 1.0, "informational": 0.55, "neutral": 0.65,
}

# top miasta PL (gazeteer lokalny) — obecność = intent lokalny (cenny dla agencji)
PL_CITIES = {
    "warszawa","warszawie","kraków","krakow","krakowie","łódź","lodz","wrocław","wroclaw",
    "poznań","poznan","gdańsk","gdansk","szczecin","bydgoszcz","lublin","białystok","bialystok",
    "katowice","katowicach","gdynia","częstochowa","czestochowa","radom","sosnowiec","toruń","torun",
    "kielce","rzeszów","rzeszow","rzeszowie","gliwice","zabrze","olsztyn","bielsko","bytom","zielona",
    "rybnik","ruda","opole","tychy","gorzów","gorzow","płock","plock","elbląg","elblag","dąbrowa",
    "wałbrzych","walbrzych","włocławek","wloclawek","tarnów","tarnow","chorzów","chorzow","koszalin",
    "kalisz","legnica","grudziądz","grudziadz","słupsk","slupsk","jaworzno","jastrzębie","nowy sącz",
    "siedlce","mysłowice","myslowice","konin","piła","pila","ostrów","ostrow","siemianowice",
    "koło","kolo","gniezno","stargard","głogów","glogow","zamość","zamosc","leszno","suwałki","suwalki",
    "pabianice","przemyśl","przemysl","ełk","elk","świdnica","swidnica","tomaszów","tomaszow","mielec",
    "śląsk","slask","pomorze","mazowsze","małopolska","malopolska","podkarpacie",
}
# słowa o intencji komercyjnej/usługowej
MONEY_WORDS = {
    "agencja","agencje","agencji","usługa","usługi","usług","cena","ceny","cennik","koszt","kosztuje",
    "ile","firma","firmy","zlecę","zlecenie","oferta","wycena","obsługa","outsourcing","specjalista",
    "konsultacja","audyt","pozycjonowanie","kampania","kampanie","reklama","reklamy","abonament",
}
# początki fraz informacyjnych / pytań
QUESTION_STARTS = ("jak ","co to","co to jest","czym jest","kiedy ","dlaczego ","czy ","gdzie ","ile ",
                   "po co","na czym","który ","jakie ","jaki ","jaka ","najlepsz","ranking","przykład",
                   "przyklady","przykłady","poradnik","wzór","wzor")
# słowa-zaślepki spinające domenę (do wycięcia przy ekstrakcji rdzenia brandu)
BRAND_FILLER = {"marketing","media","agencja","group","grupa","studio","digital","online","reklama",
                "social","seo","ads","pl","com","net","eu","www"}

# synonimy nagłówków kolumn (Senuto PL / Ahrefs / SEMrush EN) → kanon
COLMAP = {
    "keyword":     ["słowo kluczowe","slowo kluczowe","keyword","fraza","zapytanie"],
    "volume":      ["śr. mies. liczba wyszukiwań","sr. mies. liczba wyszukiwan","liczba wyszukiwań",
                    "volume","search volume","wolumen","avg. monthly searches"],
    "position":    ["pozycja","position","pos","rank"],
    "traffic":     ["szacowany ruch","estimated traffic","traffic","ruch","et"],
    "cpc":         ["cpc","koszt kliknięcia"],
    "kd":          ["trudność słowa kluczowego","trudnosc slowa kluczowego","difficulty","kd",
                    "keyword difficulty","trudność"],
    "url":         ["adres url","url","strona","page","landing page"],
}

# ───────────────────────── parsowanie wejścia ─────────────────────────

def _norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def _read_rows(path):
    """Zwraca (headers, list[dict-by-canonical-col])."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        headers = [_norm(h) for h in next(it)]
        raw = ([c for c in r] for r in it)
        rows = list(raw)
    else:
        with open(path, encoding="utf-8-sig", newline="") as f:
            # autodetekcja separatora
            sample = f.read(4096); f.seek(0)
            sep = ";" if sample.count(";") > sample.count(",") else ","
            rd = csv.reader(f, delimiter=sep)
            headers = [_norm(h) for h in next(rd)]
            rows = [r for r in rd]
    # mapowanie kolumn
    idx = {}
    for canon, names in COLMAP.items():
        for i, h in enumerate(headers):
            if any(h == n or h.startswith(n) for n in names):
                idx[canon] = i; break
    missing = [c for c in ("keyword","volume","position","url") if c not in idx]
    if missing:
        raise SystemExit(f"Nie znaleziono kolumn: {missing}\nNagłówki w pliku: {headers}")

    def num(v):
        if v is None or v == "": return 0.0
        try: return float(str(v).replace("\xa0","").replace(" ","").replace(",", "."))
        except: return 0.0

    out = []
    for r in rows:
        kw = _norm(r[idx["keyword"]])
        if not kw: continue
        out.append({
            "keyword": kw,
            "volume": int(num(r[idx["volume"]])),
            "position": num(r[idx["position"]]),
            "traffic": num(r[idx["traffic"]]) if "traffic" in idx else 0.0,
            "cpc": num(r[idx["cpc"]]) if "cpc" in idx else 0.0,
            "kd": num(r[idx["kd"]]) if "kd" in idx else 0.0,
            "url": str(r[idx["url"]] or "").strip(),
        })
    return out

# ───────────────────────── klasyfikacja fraz (reguły) ─────────────────────────

def brand_tokens(domain):
    """Z domeny wyciąga rdzeń brandu, np. kamanmarketing.pl → {'kamanmarketing','kaman'}."""
    label = re.sub(r"^www\.", "", _norm(domain)).split("/")[0].split(".")[0]
    toks = {label}
    # odetnij końcówki-zaślepki (marketing/media/agencja...) → rdzeń
    core = label
    for f in sorted(BRAND_FILLER, key=len, reverse=True):
        if core.endswith(f) and len(core) > len(f) + 2:
            core = core[: -len(f)]; break
    if core and core != label and len(core) >= 3:
        toks.add(core)
    return {t for t in toks if len(t) >= 3}

def has_nonlatin(s):
    return bool(re.search(r"[Ѐ-ӿ؀-ۿ一-鿿]", s))

def classify_kw(kw, vol, cpc, btoks):
    """Reguły → flaga intencji. 'generic_suspect' = do potwierdzenia przez Claude'a."""
    toks = kw.split()
    if has_nonlatin(kw):
        return "foreign"
    if any(b in toks or b in kw.replace(" ", "") for b in btoks):
        return "brand"
    if any(c in toks for c in PL_CITIES):
        return "local"
    if cpc >= COMMERCIAL_CPC or any(w in toks for w in MONEY_WORDS):
        return "commercial"
    # podejrzenie generyka/ambiwalentnego: 1 słowo + (duży wolumen albo brak CPC)
    if len(toks) == 1 and (vol >= GENERIC_VOL or cpc == 0):
        return "generic_suspect"
    if kw.startswith(QUESTION_STARTS):
        return "informational"
    return "neutral"

# tokeny w slugu sygnalizujące stronę ofertową/usługową (PL SEO ma długie opisowe slugi usług)
# UWAGA: tokeny celowo wąskie — np. 'realizacj' wycięte, bo łapało blogowe "omówienie realizacji".
SERVICE_SLUG_RX = re.compile(
    r"(uslugi|usluga|oferta|oferty|cennik|wynajem|sprzedaz|organizacj|dla-firm|"
    r"dla-firmy|obsluga|outsourcing|abonament)")
BLOG_SLUG_RX = re.compile(r"(^|/)(blog|poradnik|aktualnosci|artykul|wiedza|baza-wiedzy|porady|case-study)(/|$)")
TAXO_SLUG_RX = re.compile(r"(^|/)(kategoria|category|tag|tagi|produkt-kategoria)(/|$)")

# ecommerce: kody-sufiksy (Shoper/IdoSell/Presta): -cNNN kategoria, -pNNN produkt, -rNNN region/lokalizator
ECOM_CAT_RX = re.compile(r"-c\d+$")
ECOM_PROD_RX = re.compile(r"-p\d+$")
ECOM_REGION_RX = re.compile(r"-r\d+$")
CAT_PATH_RX = re.compile(r"(^|/)(kategoria|category|kategorie|c|k|dzial|sklep)(/|$)")
PROD_PATH_RX = re.compile(r"(^|/)(produkt|produkty|product|products|p|dp|towar|item)(/|$)")
# URL systemowy/parametryczny — nie jest treścią do odtworzenia
SYSTEM_RX = re.compile(r"\.(php|aspx?|jsp|cgi)(\?|$)|[?&](file|producent|sort|k_podkat|page|id|filter|szukaj|search)=")

def guess_page_type(url):
    u = url.lower().rstrip("/")
    if SYSTEM_RX.search(u):
        return "system"
    path = re.sub(r"^https?://", "", u)
    parts = path.split("/")
    segs = [s for s in parts[1:] if s]
    if not segs:
        return "homepage"
    seg_str = "/".join(segs)
    last = segs[-1]
    # ecommerce — najpierw kody-sufiksy, potem ścieżki
    if ECOM_REGION_RX.search(last):
        return "region"
    if ECOM_CAT_RX.search(last) or CAT_PATH_RX.search("/" + seg_str + "/"):
        return "category"
    if ECOM_PROD_RX.search(last) or PROD_PATH_RX.search("/" + seg_str + "/"):
        return "product"
    if TAXO_SLUG_RX.search(seg_str):
        return "taxonomy"
    if BLOG_SLUG_RX.search("/" + seg_str + "/"):
        return "blog"
    words = last.split("-")
    if any(w in PL_CITIES for w in words):
        return "service_local"
    if SERVICE_SLUG_RX.search(seg_str):
        return "service"
    # długi opisowy slug bez sygnału usługi/oferty → prawdopodobnie blog/artykuł
    if len(words) >= 5:
        return "blog"
    return "unknown"

def refine_page_type(guess, max_cpc, flag_counts):
    """Dla 'unknown'/słabego guessu dociąga typ sygnałem komercyjnym (CPC + udział fraz)."""
    qual = sum(flag_counts.get(f, 0) for f in
               ("commercial", "local", "informational", "neutral", "generic_suspect"))
    if qual == 0:
        return guess
    comm = flag_counts.get("commercial", 0) + flag_counts.get("local", 0)
    info = flag_counts.get("informational", 0)
    if guess == "unknown":
        if max_cpc >= 5 and comm / qual >= 0.3:
            return "service_local" if flag_counts.get("local", 0) >= comm * 0.5 else "service"
        if info >= comm:
            return "blog"
    # świadomie NIE przerzucamy guess=="blog" na service: poradniki z przykładami ("plan-…-przyklady")
    # bywają komercyjne, a fałszywy flip myli; finalny typ i tak ustala osąd Claude'a.
    return guess

# ───────────────────────── ingest ─────────────────────────

def cmd_ingest(args):
    rows = _read_rows(args.file)
    domain = args.domain or _norm(rows[0]["url"]).split("/")[0]
    btoks = brand_tokens(domain)

    # klasyfikacja per-fraza
    for r in rows:
        r["flag"] = classify_kw(r["keyword"], r["volume"], r["cpc"], btoks)
        r["value"] = r["volume"] * INTENT_WEIGHT.get(r["flag"], 0.6)

    # agregacja po URL
    pages = {}
    for r in rows:
        u = r["url"]
        p = pages.setdefault(u, {
            "url": u, "kws": [], "raw_traffic": 0.0, "qualified_traffic": 0.0,
            "total_vol": 0, "max_cpc": 0.0, "kd_sum": 0.0, "kd_n": 0, "best_pos": 999,
        })
        p["kws"].append(r)
        p["raw_traffic"] += r["traffic"]
        if r["flag"] not in ("brand", "foreign"):
            p["qualified_traffic"] += r["traffic"]
        p["total_vol"] += r["volume"]
        p["max_cpc"] = max(p["max_cpc"], r["cpc"])
        if r["kd"]:
            p["kd_sum"] += r["kd"]; p["kd_n"] += 1
        p["best_pos"] = min(p["best_pos"], r["position"])

    page_list = []
    for u, p in pages.items():
        qkws = [k for k in p["kws"] if k["flag"] not in ("brand", "foreign")]
        non_suspect = [k for k in qkws if k["flag"] != "generic_suspect"]
        # main keyword = najwyższy wolumen wśród "dobrych"; fallback: dowolny qualified
        pool = non_suspect or qkws or p["kws"]
        main_by_vol = max(pool, key=lambda k: k["volume"])
        main_by_traffic = max(p["kws"], key=lambda k: k["traffic"])
        fc = defaultdict(int)
        for k in p["kws"]:
            fc[k["flag"]] += 1
        pt = refine_page_type(guess_page_type(u), p["max_cpc"], fc)
        # winsoryzacja CPC: max_cpc liczony z odrzuceniem błędów danych (>CPC_WINSOR)
        cpc_w = max((min(k["cpc"], CPC_WINSOR) for k in p["kws"]), default=0.0)
        cpc_outlier = p["max_cpc"] > CPC_WINSOR
        # latent_potential: niezagospodarowany popyt = wolumen istotnych fraz, gdzie konkurent słabo rankuje
        latent = sum(k["volume"] for k in qkws
                     if k["position"] >= WEAK_POS_MIN and k["volume"] >= LATENT_MIN_VOL)
        page_list.append({
            "url": u,
            "page_type_guess": pt,
            "raw_traffic": round(p["raw_traffic"], 1),
            "qualified_traffic": round(p["qualified_traffic"], 1),
            "kw_count": len(p["kws"]),
            "qualified_kw_count": len(qkws),
            "total_vol": p["total_vol"],
            "max_cpc": round(cpc_w, 2),
            "max_cpc_raw": round(p["max_cpc"], 2),
            "cpc_outlier": cpc_outlier,
            "latent_potential": latent,
            "open_goal": latent >= OPEN_GOAL_MIN,
            "avg_kd": round(p["kd_sum"] / p["kd_n"], 1) if p["kd_n"] else 0.0,
            "best_position": p["best_pos"],
            "main_kw_by_vol": main_by_vol["keyword"],
            "main_kw_by_traffic": main_by_traffic["keyword"],
            "keywords": sorted(
                [{"kw": k["keyword"], "vol": k["volume"], "pos": k["position"],
                  "cpc": k["cpc"], "traffic": round(k["traffic"], 1), "flag": k["flag"]}
                 for k in p["kws"]],
                key=lambda x: -x["traffic"]),
        })

    page_list.sort(key=lambda p: -p["qualified_traffic"])
    os.makedirs(args.out, exist_ok=True)

    with open(os.path.join(args.out, "pages.json"), "w", encoding="utf-8") as f:
        json.dump({"domain": domain, "brand_tokens": sorted(btoks),
                   "our_domain": args.our_domain, "our_topics": args.our_topics,
                   "pages": page_list}, f, ensure_ascii=False, indent=2)

    with open(os.path.join(args.out, "keywords.csv"), "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["keyword","volume","position","traffic","cpc","kd","flag","url"])
        for r in sorted(rows, key=lambda x: -x["traffic"]):
            w.writerow([r["keyword"], r["volume"], r["position"], round(r["traffic"],1),
                        r["cpc"], r["kd"], r["flag"], r["url"]])

    _write_worksheet(args.out, domain, btoks, args.our_domain, args.our_topics, page_list, args.top)

    # podsumowanie na stdout
    flag_counts = defaultdict(int)
    for r in rows: flag_counts[r["flag"]] += 1
    print(f"OK. domena={domain}  brand={sorted(btoks)}  fraz={len(rows)}  stron={len(page_list)}")
    print("flagi fraz:", dict(flag_counts))
    print(f"→ {args.out}/pages.json, keywords.csv, worksheet.md (osądź WSZYSTKIE strony z worksheet)")
    print("NASTĘPNY KROK: przeczytaj worksheet.md i wypełnij judgments.json (patrz SKILL.md krok 2).")

def _select_worksheet_pages(pages, top):
    """3 pule, by osąd objął nie tylko strony z dużym OBECNYM ruchem:
       (📊) top wg ruchu · (💰) top wg CPC/komercji · (🎯) top wg latent (open goals).
       Zwraca listę stron z adnotacją `pools` (które pule je wyłoniły)."""
    n_cpc = max(12, top // 3); n_open = max(18, top // 2)
    by_traffic = sorted(pages, key=lambda p: -p["qualified_traffic"])[:top]
    by_cpc = sorted([p for p in pages if p["max_cpc"] >= COMMERCIAL_CPC],
                    key=lambda p: -p["max_cpc"])[:n_cpc]
    by_open = sorted([p for p in pages if p["open_goal"]],
                     key=lambda p: -p["latent_potential"])[:n_open]
    pools = {}
    for p in by_traffic: pools.setdefault(p["url"], set()).add("📊ruch")
    for p in by_cpc:     pools.setdefault(p["url"], set()).add("💰cpc")
    for p in by_open:    pools.setdefault(p["url"], set()).add("🎯open")
    sel = [p for p in pages if p["url"] in pools]
    for p in sel: p["_pools"] = pools[p["url"]]
    # sort: najpierw ruch, potem latent — żeby czytało się sensownie
    sel.sort(key=lambda p: (-p["qualified_traffic"], -p["latent_potential"]))
    return sel

def _write_worksheet(out, domain, btoks, our_domain, our_topics, pages, top):
    sel = _select_worksheet_pages(pages, top)
    L = []
    nf = our_topics or our_domain
    L.append(f"# Worksheet osądu — {domain}")
    L.append("")
    L.append(f"Brand: `{', '.join(sorted(btoks))}`"
             + (f" · Filtr niszy: `{nf}`" if nf else " · Tryb: ranking wg wartości strony do odtworzenia"))
    L.append("")
    L.append(f"**Osądź KAŻDĄ z {len(sel)} stron poniżej** → wpisz do `judgments.json` (klucz = url). "
             "Pola: `page_type`, `main_keyword`, `relevance` (0–1 wartość strony DO ODTWORZENIA"
             + (" + dopasowanie do podanej niszy" if nf else "")
             + "), `junk_keywords`, `content_type`, `note`. Szczegóły w SKILL.md. "
             "Strony NIE osądzone trafią do długiego ogona (poza P1/P2/P3) — dlatego nie pomijaj.")
    L.append("")
    L.append("Pule selekcji: **📊ruch** (duży obecny ruch) · **💰cpc** (wysoka wartość komercyjna) · "
             "**🎯open** (open goal: jest popyt, konkurent słabo rankuje → łatwy łup). "
             "Strona z 🎯 ale bez 📊 = mały ruch DZIŚ, ale duży potencjał.")
    L.append("")
    L.append("Flagi reguł: brand/foreign = wykluczone z ruchu · generic_suspect = "
             "**TY potwierdzasz** czy bezwartościowe/ambiwalentne (np. „landing\", „biogram\"). "
             "⚠cpc = CPC zawyżone (błąd danych Senuto) — nie sugeruj się.")
    L.append("")
    for i, p in enumerate(sel, 1):
        cpc_txt = f"{p['max_cpc']:g}" + (f" ⚠(raw {p['max_cpc_raw']:g})" if p.get("cpc_outlier") else "")
        L.append(f"## {i}. {p['url']}")
        L.append(f"- typ: **{p['page_type_guess']}** · pule: {' '.join(sorted(p.get('_pools', [])))} · "
                 f"ruch_now: **{p['qualified_traffic']}** · latent(open): {p['latent_potential']} · "
                 f"fraz: {p['kw_count']} (qual {p['qualified_kw_count']}) · vol_suma: {p['total_vol']} · "
                 f"max_cpc: {cpc_txt} · avg_kd: {p['avg_kd']} · best_pos: {p['best_position']:g}")
        L.append(f"- main_kw kandydaci → wg wolumenu: `{p['main_kw_by_vol']}` · "
                 f"wg ruchu: `{p['main_kw_by_traffic']}`")
        L.append("- TOP frazy (kw · vol · pos · cpc · ruch · flaga):")
        for k in p["keywords"][:9]:
            warn = " ⚠cpc" if k["cpc"] > CPC_WINSOR else ""
            L.append(f"    - {k['kw']} · {k['vol']} · poz {k['pos']:g} · "
                     f"cpc {k['cpc']:g}{warn} · ruch {k['traffic']:g} · [{k['flag']}]")
        L.append("")
    with open(os.path.join(out, "worksheet.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(L))

# ───────────────────────── score ─────────────────────────

def cmd_score(args):
    with open(os.path.join(args.inp, "pages.json"), encoding="utf-8") as f:
        data = json.load(f)
    with open(args.judgments, encoding="utf-8") as f:
        judg = json.load(f)

    pages = data["pages"]
    # przelicz effective_traffic ORAZ effective_latent po wykluczeniu junk z osądu Claude'a
    for p in pages:
        j = judg.get(p["url"], {})
        p["judged"] = p["url"] in judg
        junk = set(_norm(x) for x in j.get("junk_keywords", []))
        nonjunk = [k for k in p["keywords"]
                   if k["flag"] not in ("brand", "foreign") and _norm(k["kw"]) not in junk]
        p["effective_traffic"] = round(sum(k["traffic"] for k in nonjunk), 1)
        # latent PO usunięciu junk — inaczej generyk-śmieć (np. „atrakcji" 450k) zawyża open-goals
        p["latent_potential"] = sum(k["vol"] for k in nonjunk
                                    if k["pos"] >= WEAK_POS_MIN and k["vol"] >= LATENT_MIN_VOL)
        p["open_goal"] = p["latent_potential"] >= OPEN_GOAL_MIN
        p["relevance"] = float(j.get("relevance", 0.6))
        p["final_type"] = j.get("page_type", p["page_type_guess"])
        p["final_main_kw"] = j.get("main_keyword", p["main_kw_by_vol"])
        p["content_type"] = j.get("content_type", p["final_type"])
        p["note"] = j.get("note", "")
        # opportunity = ruch DZIŚ + odzyskiwalny popyt latentny (żeby open-goale z zerowym
        # obecnym ruchem, ale dużym wolumenem na słabych pozycjach, też trafiały do P1/P2)
        p["opportunity"] = p["effective_traffic"] + LATENT_W * p["latent_potential"]

    max_opp = max((p["opportunity"] for p in pages), default=1) or 1
    max_kw = max((p["qualified_kw_count"] for p in pages), default=1) or 1

    for p in pages:
        # sqrt: malejące zwroty — jeden outlier nie spłaszcza reszty
        ntraf = math.sqrt(p["opportunity"]) / math.sqrt(max_opp)
        nbreadth = math.log1p(p["qualified_kw_count"]) / math.log1p(max_kw)
        ncomm = min(p["max_cpc"] / CPC_CAP, 1.0)   # max_cpc jest już winsoryzowane (ingest)
        kd_pen = (p["avg_kd"] / 100.0) * KD_PENALTY
        base = (W_TRAFFIC * ntraf + W_BREADTH * nbreadth + W_COMMERCIAL * ncomm)
        p["score"] = round(p["relevance"] * base * (1 - kd_pen) * 100, 1)

    # ── Tiering ──
    # Nieosądzone strony NIE są rekomendacjami (inaczej brand/generyk z default relevance 0.6
    # przeciekają do P3). Tiery P1/P2/P3 są WZGLĘDNE wśród osądzonych, kwalifikujących się stron —
    # dzięki temu "category = P1" działa też dla małego konkurenta (ranking, nie próg absolutny).
    eligible = []
    for p in pages:
        if p["final_type"] == "system":
            p["tier"] = "SKIP"          # URL parametryczny — nigdy nie jest treścią do odtworzenia
        elif not p["judged"]:
            p["tier"] = "TAIL"          # długi ogon — nieanalizowane, poza rekomendacjami
        elif p["relevance"] < RELEVANCE_SKIP:
            p["tier"] = "SKIP"          # świadomie odpuszczone (brand/off-topic)
        else:
            eligible.append(p)
    # twardy floor: relevance < FLOOR → max P3 (traffic-bait nie wejdzie do P1/P2 mimo ruchu)
    for p in [p for p in eligible if p["relevance"] < RELEVANCE_FLOOR_P3]:
        p["tier"] = "P3"
    ranked = sorted([p for p in eligible if p["relevance"] >= RELEVANCE_FLOOR_P3],
                    key=lambda p: -p["score"])
    nr = len(ranked)
    for i, p in enumerate(ranked):
        pct = (i + 0.5) / nr if nr else 1.0
        p["tier"] = "P1" if pct <= TIER_P1_PCT else "P2" if pct <= TIER_P2_PCT else "P3"

    pages.sort(key=lambda p: (-p["score"]))
    out = args.out or args.inp
    _write_report(out, data, pages)
    with open(os.path.join(out, "priorities.csv"), "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["tier","score","url","final_type","main_keyword","effective_traffic",
                    "latent_potential","open_goal","qualified_kw_count","max_cpc","avg_kd",
                    "best_position","relevance","judged","content_type","note"])
        for p in pages:
            w.writerow([p["tier"], p["score"], p["url"], p["final_type"], p["final_main_kw"],
                        p["effective_traffic"], p["latent_potential"], int(p["open_goal"]),
                        p["qualified_kw_count"], p["max_cpc"], p["avg_kd"], p["best_position"],
                        p["relevance"], int(p["judged"]), p["content_type"], p["note"]])
    tiers = defaultdict(int)
    for p in pages: tiers[p["tier"]] += 1
    print(f"OK. {dict(tiers)} → {out}/report.md, priorities.csv")
    print("   (TAIL = nieosądzone, poza rekomendacjami; osądź więcej stron, by je podnieść)")

def _write_report(out, data, pages):
    L = [f"# Plan odtworzenia ruchu — {data['domain']}", ""]
    nf = data.get("our_topics") or data.get("our_domain")
    L.append(f"Brand: `{', '.join(data['brand_tokens'])}`"
             + (f" · Filtr niszy: `{nf}`" if nf else " · Tryb: ranking wg wartości strony do odtworzenia"))
    L.append("")
    L.append("Priorytet = wartość strony × (√(ruch+latent) + szerokość fraz + wartość komercyjna) "
             "− kara za trudność. Tiery są **względne** (ranking osądzonych stron): "
             "**P1** = top, rób najpierw · **P2/P3** niżej · **SKIP** = świadomie odpuszczone "
             "(brand/off-topic/system) · **🎯** = open goal (jest popyt, konkurent słabo rankuje → łatwy łup).")
    L.append("")

    # rozkład typów stron — pomaga zauważyć LUKI (np. konkurent nie ma bloga → niezagospodarowany popyt info)
    from collections import Counter
    tdist = Counter(p["final_type"] for p in pages if p["tier"] != "TAIL")
    if tdist:
        L.append("**Typy stron (osądzone):** "
                 + " · ".join(f"{t}={n}" for t, n in tdist.most_common()))
        L.append("")

    for tier in ("P1", "P2", "P3", "SKIP"):
        grp = [p for p in pages if p["tier"] == tier]
        if not grp: continue
        L.append(f"## {tier} ({len(grp)})")
        L.append("")
        for p in grp:
            goal = " 🎯" if p.get("open_goal") and p["tier"] != "SKIP" else ""
            L.append(f"### [{p['score']}]{goal} {p['final_main_kw']}  ·  _{p['final_type']}_")
            L.append(f"- **Źródło konkurenta:** {p['url']}")
            L.append(f"- **Co zbudować:** {p['content_type']}"
                     + (f" — {p['note']}" if p["note"] else ""))
            pot = (f" · **potencjał (open goal):** {p['latent_potential']} wol. na słabo zajętych frazach"
                   if p.get("open_goal") else "")
            L.append(f"- **Ruch do przejęcia:** {p['effective_traffic']} / mc "
                     f"· fraz: {p['qualified_kw_count']} · max CPC: {p['max_cpc']} "
                     f"· trudność (KD): {p['avg_kd']} · konkurent na poz. {p['best_position']:g}{pot}")
            support = [k["kw"] for k in p["keywords"]
                       if k["flag"] not in ("brand", "foreign") and k["kw"] != p["final_main_kw"]][:6]
            if support:
                L.append(f"- **Frazy wspierające:** {', '.join(support)}")
            L.append("")

    # Open goals — tylko rekomendowane (P1/P2): duży latentny popyt + dopasowanie do nas.
    # P3 świadomie pomijamy — open-goal z niskim relevance to nie "szybki łup", tylko szum.
    goals = sorted([p for p in pages if p.get("open_goal") and p["tier"] in ("P1", "P2")],
                   key=lambda p: -p["latent_potential"])[:10]
    if goals:
        L.append("## 🎯 Open goals — szybkie łupy (jest popyt, konkurent słabo rankuje)")
        L.append("")
        for p in goals:
            L.append(f"- **{p['final_main_kw']}** ({p['final_type']}, {p['tier']}) — "
                     f"latent {p['latent_potential']} wol., konkurent poz. {p['best_position']:g}, "
                     f"obecny ruch {p['effective_traffic']}/mc · {p['url']}")
        L.append("")

    # Długi ogon — nieosądzone (NIE rekomendacje; kandydaci do pogłębienia jeśli mają ruch/potencjał)
    tail = [p for p in pages if p["tier"] == "TAIL"]
    if tail:
        tail_sorted = sorted(tail, key=lambda p: -(p["effective_traffic"] + p["latent_potential"] * 0.1))
        shown = [p for p in tail_sorted if p["effective_traffic"] > 0 or p["open_goal"]][:12]
        L.append(f"## Długi ogon — nieosądzone ({len(tail)})")
        L.append("")
        L.append("Strony spoza puli osądu — **niezweryfikowane**, nie traktuj jako rekomendacji. "
                 "Jeśli któraś poniżej ma realny ruch/potencjał, dorzuć ją do `judgments.json` i przelicz `score`.")
        L.append("")
        for p in shown:
            L.append(f"- _{p['final_type']}_ · ruch {p['effective_traffic']}/mc · "
                     f"latent {p['latent_potential']} · {p['url']}")
        L.append("")

    with open(os.path.join(out, "report.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(L))

# ───────────────────────── CLI ─────────────────────────

def main():
    ap = argparse.ArgumentParser(prog="recon_tools")
    sub = ap.add_subparsers(dest="cmd", required=True)

    pi = sub.add_parser("ingest")
    pi.add_argument("file")
    pi.add_argument("--domain", default=None, help="domena konkurenta (auto z URL jeśli brak)")
    pi.add_argument("--our-domain", default=None, help="OPCJONALNIE: domena, którą rozwijamy — gdy chcesz dodatkowo filtrować pod własną stronę")
    pi.add_argument("--our-topics", default=None, help="OPCJONALNIE: opis niszy/tematów do filtra dopasowania (np. 'baterie łazienkowe'); bez tego skill rankuje wg wartości strony do odtworzenia")
    pi.add_argument("--top", type=int, default=40, help="ile stron w worksheet do osądu")
    pi.add_argument("--out", default="recon_out")
    pi.set_defaults(func=cmd_ingest)

    ps = sub.add_parser("score")
    ps.add_argument("--in", dest="inp", required=True)
    ps.add_argument("--judgments", required=True)
    ps.add_argument("--out", default=None)
    ps.set_defaults(func=cmd_score)

    args = ap.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()
